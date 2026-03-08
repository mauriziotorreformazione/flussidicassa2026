"""
Generazione file Excel Piano dei Flussi di Cassa.
Produce 3 fogli: ENTRATE, SPESE, PIANO FLUSSI (formato MIM).
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── COLORS ──────────────────────────────────────────────
COLOR_HEADER_BLUE = "1F4E79"       # Dark blue header
COLOR_HEADER_LIGHT = "BDD7EE"      # Light blue subheader
COLOR_ENTRATE = "E2EFDA"           # Light green - entrate
COLOR_SPESE = "FCE4D6"             # Light orange - spese
COLOR_RESIDUI = "FFF2CC"           # Light yellow - residui
COLOR_TOTALE = "D6DCE4"            # Grey - totals
COLOR_ANOMALIA = "FFFF00"          # Orange - voci programmaz=0 con movimenti
COLOR_PIANO_HEADER = "203864"      # Very dark blue for piano flussi
COLOR_PIANO_INCASSO = "DEEAF1"     # Light for piano flussi rows

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_SUBHEADER = Font(name="Calibri", bold=True, size=10)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_ANOMALIA = Font(name="Calibri", bold=True, color="C00000", size=10)

BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
BORDER_MEDIUM = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium")
)

NUM_EURO = '#,##0.00 €'
NUM_PERC = '0%'


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _header_row(ws, row, values, fill_color, font=None, height=18):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(fill_color)
        cell.font = font or FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def _data_row(ws, row, values, fill_color=None, font=None, formats=None, anomalia=False):
    ws.row_dimensions[row].height = 16
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if fill_color:
            cell.fill = _fill(fill_color)
        if anomalia:
            cell.font = FONT_ANOMALIA
        else:
            cell.font = font or FONT_NORMAL
        cell.border = BORDER_THIN
        if formats and col - 1 < len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]
        # Align numbers right
        if isinstance(val, (int, float)):
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def calcola_flussi(voce: dict, percentuali: dict, tipo: str, modalita: str) -> tuple:
    """
    Calcola Gen-Ago e Set-Dic per una voce.
    tipo: 'entrata_comp' | 'spesa_comp' | 'residuo_attivo' | 'residuo_passivo'
    Restituisce (gen_ago, set_dic)
    """
    if tipo == "entrata_comp":
        programmaz = voce.get("previsione_definitiva", 0.0)
        gia_incassato = voce.get("somme_riscosse", 0.0)
        pct_ga = percentuali.get("entrate_competenza_gen_ago", 100) / 100
        pct_sd = percentuali.get("entrate_competenza_set_dic", 0) / 100
    elif tipo == "spesa_comp":
        programmaz = voce.get("previsione_definitiva", 0.0)
        gia_incassato = voce.get("somme_pagate", 0.0)
        pct_ga = percentuali.get("spese_competenza_gen_ago", 67) / 100
        pct_sd = percentuali.get("spese_competenza_set_dic", 33) / 100
    elif tipo == "residuo_attivo":
        programmaz = voce.get("importo", 0.0)
        gia_incassato = 0.0
        pct_ga = percentuali.get("residui_attivi_gen_ago", 33) / 100
        pct_sd = percentuali.get("residui_attivi_set_dic", 67) / 100
    else:  # residuo_passivo
        programmaz = voce.get("importo", 0.0)
        gia_incassato = 0.0
        pct_ga = percentuali.get("residui_passivi_gen_ago", 33) / 100
        pct_sd = percentuali.get("residui_passivi_set_dic", 67) / 100

    prog_originale_zero = voce.get("prog_zero", False)
    anomalia_zero = prog_originale_zero and gia_incassato > 0

    if anomalia_zero:
        return gia_incassato, 0.0, True  # anomalia gialla

    # Per entrate: se programmaz < riscosso, allinea programmaz al riscosso (anomalia gialla)
    if tipo in ("entrata_comp", "residuo_attivo") and gia_incassato > programmaz:
        programmaz = gia_incassato  # allineamento silenzioso, segnalato con flag
        return gia_incassato, 0.0, True  # anomalia gialla

    differenza = programmaz - gia_incassato
    if differenza < 0:
        differenza = 0.0

    gen_ago = round(gia_incassato + differenza * pct_ga, 2)
    set_dic = round(differenza * pct_sd, 2)
    return gen_ago, set_dic, False


def crea_foglio_entrate(wb, dati, percentuali, modalita, minute_spese=0.0):
    ws = wb.create_sheet("ENTRATE")
    _set_col_widths(ws, {"A": 12, "B": 45, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    row = 1
    # Title
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="PIANO DEI FLUSSI DI CASSA — ENTRATE")
    cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    cell.fill = _fill(COLOR_HEADER_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

    row += 1
    headers = ["Codice", "Descrizione voce", "Programm. Definitiva", "Già Riscosso",
               "Gen–Ago", "Set–Dic", "TOTALE"]
    _header_row(ws, row, headers, COLOR_HEADER_BLUE)

    fmt = [None, None, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO]

    # ── SEZIONE 1: Entrate in competenza ──────────────────
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="ENTRATE IN COMPETENZA (Modello H)")
    cell.font = FONT_SUBHEADER
    cell.fill = _fill(COLOR_ENTRATE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 18

    totale_prog_comp = 0
    totale_gia_comp = 0
    totale_ga_comp = 0
    totale_sd_comp = 0

    for voce in dati.get("entrate", []):
        codice = voce.get("codice", "")
        # Skip E1/1 and E1/2 (avanzo di amministrazione - excluded)
        if codice in ["E1/1", "E1/2"]:
            continue

        row += 1
        prog = voce.get("previsione_definitiva", 0.0)
        gia = voce.get("somme_riscosse", 0.0)

        if modalita == "automatica":
            gen_ago, set_dic, anomalia = calcola_flussi(voce, percentuali, "entrata_comp", modalita)
        else:
            gen_ago = gia
            set_dic = 0.0
            anomalia = (prog == 0 and gia > 0) or (gia > prog > 0)

        # Se riscosso > programmaz, mostra programmaz allineata al riscosso
        prog_display = max(prog, gia) if gia > prog else prog

        totale_row = gen_ago + set_dic

        _data_row(ws, row,
                  [codice, voce.get("descrizione", ""), prog_display, gia, gen_ago, set_dic, totale_row],
                  fill_color=COLOR_ANOMALIA if anomalia else None,
                  formats=fmt, anomalia=anomalia)

        totale_prog_comp += prog
        totale_gia_comp += gia
        totale_ga_comp += gen_ago
        totale_sd_comp += set_dic

    # Subtotal entrate competenza
    row += 1
    _data_row(ws, row,
              ["", "TOTALE ENTRATE COMPETENZA",
               totale_prog_comp, totale_gia_comp,
               totale_ga_comp, totale_sd_comp,
               totale_ga_comp + totale_sd_comp],
              fill_color=COLOR_TOTALE, font=FONT_BOLD, formats=fmt)

    # ── SEZIONE 2: Residui attivi ──────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="RESIDUI ATTIVI — crediti anni precedenti (Modello L)")
    cell.font = FONT_SUBHEADER
    cell.fill = _fill(COLOR_RESIDUI)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 18

    totale_prog_res = 0
    totale_ga_res = 0
    totale_sd_res = 0

    for residuo in dati.get("residui_attivi", []):
        row += 1
        importo = residuo.get("importo", 0.0)

        if modalita == "automatica":
            gen_ago, set_dic, anomalia = calcola_flussi(residuo, percentuali, "residuo_attivo", modalita)
        else:
            gen_ago = 0.0
            set_dic = 0.0
            anomalia = False

        desc = f"{residuo.get('anno', '')} - {residuo.get('oggetto', residuo.get('debitore', ''))}"
        _data_row(ws, row,
                  [residuo.get("codice_pdc", ""), desc, importo, 0.0, gen_ago, set_dic, gen_ago + set_dic],
                  fill_color=COLOR_RESIDUI if not anomalia else COLOR_ANOMALIA,
                  formats=fmt, anomalia=anomalia)

        totale_prog_res += importo
        totale_ga_res += gen_ago
        totale_sd_res += set_dic

    # Subtotal residui attivi
    row += 1
    _data_row(ws, row,
              ["", "TOTALE RESIDUI ATTIVI",
               totale_prog_res, 0.0,
               totale_ga_res, totale_sd_res,
               totale_ga_res + totale_sd_res],
              fill_color=COLOR_TOTALE, font=FONT_BOLD, formats=fmt)

    # ── GRAND TOTAL ENTRATE ────────────────────────────────
    row += 2
    gt_vals = [
        "TOTALE GENERALE ENTRATE",
        totale_prog_comp + totale_prog_res,
        totale_gia_comp,
        totale_ga_comp + totale_ga_res,
        totale_sd_comp + totale_sd_res,
        totale_ga_comp + totale_ga_res + totale_sd_comp + totale_sd_res,
    ]
    gt_cols = [1, 3, 4, 5, 6, 7]
    gt_fmts = [None, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO]
    ws.row_dimensions[row].height = 18
    for idx, (col, val) in enumerate(zip(gt_cols, gt_vals)):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(COLOR_HEADER_BLUE)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.border = BORDER_THIN
        cell.number_format = gt_fmts[idx] or "General"
        cell.alignment = Alignment(horizontal="right" if isinstance(val, float) else "left", vertical="center")
    # Merge col 1-2 for label
    ws.merge_cells(f"A{row}:B{row}")

    # ── PARTITE DI GIRO: Minute Spese ────────────────────────
    if minute_spese > 0:
        row += 1
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1, value="PARTITE DI GIRO")
        cell.font = FONT_SUBHEADER
        cell.fill = _fill("E2CFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER_THIN
        ws.row_dimensions[row].height = 16
        row += 1
        _data_row(ws, row,
                  ["E.9.01", "FONDO MINUTE SPESE (partita di giro — ricostituzione)",
                   minute_spese, 0.0, 0.0, minute_spese, minute_spese],
                  fill_color="E2CFFF", formats=fmt)

    if modalita == "manuale":
        row += 2
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1,
                       value="⚠️ MODALITÀ MANUALE: inserire i valori Gen-Ago e Set-Dic per ogni voce. "
                             "Gen-Ago non può essere inferiore al 'Già Riscosso'. "
                             "Le celle evidenziate in giallo hanno Programmazione=0 con movimenti.")
        cell.font = Font(name="Calibri", bold=True, color="C00000", size=10)
        cell.fill = _fill("FFFFC0")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 32

    return {
        "totale_entrate_ga": totale_ga_comp + totale_ga_res,
        "totale_entrate_sd": totale_sd_comp + totale_sd_res + minute_spese,
    }


def crea_foglio_spese(wb, dati, percentuali, modalita, minute_spese=0.0):
    ws = wb.create_sheet("SPESE")
    _set_col_widths(ws, {"A": 12, "B": 45, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="PIANO DEI FLUSSI DI CASSA — SPESE")
    cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    cell.fill = _fill("C00000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

    row += 1
    headers = ["Codice", "Descrizione voce", "Programm. Definitiva", "Già Pagato",
               "Gen–Ago", "Set–Dic", "TOTALE"]
    _header_row(ws, row, headers, "C00000")

    fmt = [None, None, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO]

    # ── SEZIONE 1: Spese in competenza ────────────────────
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="SPESE IN COMPETENZA (Modello H)")
    cell.font = FONT_SUBHEADER
    cell.fill = _fill(COLOR_SPESE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 18

    totale_prog_comp = 0
    totale_gia_comp = 0
    totale_ga_comp = 0
    totale_sd_comp = 0

    for voce in dati.get("spese", []):
        row += 1
        prog = voce.get("previsione_definitiva", 0.0)
        gia = voce.get("somme_pagate", 0.0)

        if modalita == "automatica":
            gen_ago, set_dic, anomalia = calcola_flussi(voce, percentuali, "spesa_comp", modalita)
        else:
            gen_ago = gia
            set_dic = 0.0
            anomalia = (prog == 0 and gia > 0)

        _data_row(ws, row,
                  [voce.get("aggregato", ""), voce.get("descrizione", ""),
                   prog, gia, gen_ago, set_dic, gen_ago + set_dic],
                  fill_color=COLOR_ANOMALIA if anomalia else None,
                  formats=fmt, anomalia=anomalia)

        totale_prog_comp += prog
        totale_gia_comp += gia
        totale_ga_comp += gen_ago
        totale_sd_comp += set_dic

    row += 1
    _data_row(ws, row,
              ["", "TOTALE SPESE COMPETENZA",
               totale_prog_comp, totale_gia_comp,
               totale_ga_comp, totale_sd_comp,
               totale_ga_comp + totale_sd_comp],
              fill_color=COLOR_TOTALE, font=FONT_BOLD, formats=fmt)

    # ── SEZIONE 2: Residui passivi ─────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="RESIDUI PASSIVI — debiti anni precedenti (Modello L)")
    cell.font = FONT_SUBHEADER
    cell.fill = _fill(COLOR_RESIDUI)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 18

    totale_prog_res = 0
    totale_ga_res = 0
    totale_sd_res = 0

    for residuo in dati.get("residui_passivi", []):
        row += 1
        importo = residuo.get("importo", 0.0)

        if modalita == "automatica":
            gen_ago, set_dic, anomalia = calcola_flussi(residuo, percentuali, "residuo_passivo", modalita)
        else:
            gen_ago = 0.0
            set_dic = 0.0
            anomalia = False

        desc = f"{residuo.get('anno', '')} - {residuo.get('oggetto', residuo.get('creditore', ''))}"
        _data_row(ws, row,
                  [residuo.get("codice_pdc", ""), desc, importo, 0.0, gen_ago, set_dic, gen_ago + set_dic],
                  fill_color=COLOR_RESIDUI if not anomalia else COLOR_ANOMALIA,
                  formats=fmt, anomalia=anomalia)

        totale_prog_res += importo
        totale_ga_res += gen_ago
        totale_sd_res += set_dic

    row += 1
    _data_row(ws, row,
              ["", "TOTALE RESIDUI PASSIVI",
               totale_prog_res, 0.0,
               totale_ga_res, totale_sd_res,
               totale_ga_res + totale_sd_res],
              fill_color=COLOR_TOTALE, font=FONT_BOLD, formats=fmt)

    row += 2
    gt_vals = [
        "TOTALE GENERALE SPESE",
        totale_prog_comp + totale_prog_res,
        totale_gia_comp,
        totale_ga_comp + totale_ga_res,
        totale_sd_comp + totale_sd_res,
        totale_ga_comp + totale_ga_res + totale_sd_comp + totale_sd_res,
    ]
    gt_cols = [1, 3, 4, 5, 6, 7]
    gt_fmts = [None, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO]
    ws.row_dimensions[row].height = 18
    for idx, (col, val) in enumerate(zip(gt_cols, gt_vals)):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill("C00000")
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.border = BORDER_THIN
        cell.number_format = gt_fmts[idx] or "General"
        cell.alignment = Alignment(horizontal="right" if isinstance(val, float) else "left", vertical="center")
    ws.merge_cells(f"A{row}:B{row}")

    # ── PARTITE DI GIRO: Minute Spese ────────────────────────
    if minute_spese > 0:
        row += 1
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1, value="PARTITE DI GIRO")
        cell.font = FONT_SUBHEADER
        cell.fill = _fill("E2CFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER_THIN
        ws.row_dimensions[row].height = 16
        row += 1
        _data_row(ws, row,
                  ["U.7.01", "FONDO MINUTE SPESE (partita di giro — utilizzo)",
                   minute_spese, 0.0, minute_spese, 0.0, minute_spese],
                  fill_color="E2CFFF", formats=fmt)

    if modalita == "manuale":
        row += 2
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1,
                       value="⚠️ MODALITÀ MANUALE: inserire i valori Gen-Ago e Set-Dic per ogni voce.")
        cell.font = Font(name="Calibri", bold=True, color="C00000", size=10)
        cell.fill = _fill("FFFFC0")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    return {
        "totale_spese_ga": totale_ga_comp + totale_ga_res + minute_spese,
        "totale_spese_sd": totale_sd_comp + totale_sd_res,
    }


def crea_foglio_piano_flussi(wb, dati_scuola, totali, fondo_cassa):
    """Foglio Piano Flussi con struttura ministeriale ufficiale (Allegato 3 MIM 2284/2025)."""
    ws = wb.create_sheet("PIANO FLUSSI (MIM)")
    _set_col_widths(ws, {"A": 8, "B": 55, "C": 22, "D": 22, "E": 22, "F": 22})

    anno = dati_scuola.get("anno_esercizio", "2026")
    nome = dati_scuola.get("nome_istituto", "ISTITUTO")
    minute_spese = float(dati_scuola.get("fondo_minute_spese", 0))

    # Totali dai fogli 1 e 2
    ent_ga = totali.get("totale_entrate_ga", 0.0)
    ent_sd = totali.get("totale_entrate_sd", 0.0)
    spe_ga = totali.get("totale_spese_ga", 0.0)
    spe_sd = totali.get("totale_spese_sd", 0.0)

    # Totali per codice PDC aggregato
    pdc = totali.get("pdc", {})

    def pdc_ga(code): return pdc.get(code, {}).get("ga", 0.0)
    def pdc_sd(code): return pdc.get(code, {}).get("sd", 0.0)
    def pdc_tot(code): return pdc_ga(code) + pdc_sd(code)

    row = 1

    def title_row(text, color=COLOR_PIANO_HEADER, size=11, height=20):
        nonlocal row
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Calibri", bold=True, size=size, color="FFFFFF")
        c.fill = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = height
        row += 1

    def header_cols():
        nonlocal row
        headers = [
            "Cod. PDC",
            "Descrizione",
            f"Totale anno (12/12)\nPrevisioni di cassa",
            f"Al 31/08 (1/1–31/8)\nPrevisioni di cassa",
            f"Ultimi 4/12 (1/9–31/12)\nDi cui 4/12",
            f"Al 31/12 (cumulato)\nPrevisioni di cassa",
        ]
        ws.row_dimensions[row].height = 30
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            c.fill = _fill(COLOR_PIANO_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER_THIN
        row += 1

    def data_row_mim(codice, desc, totale_12, al_31ago, di_cui_4, al_31dic,
                     fill=None, bold=False, indent=0):
        nonlocal row
        desc_full = ("    " * indent) + desc
        vals = [codice, desc_full, totale_12, al_31ago, di_cui_4, al_31dic]
        fmts = [None, None, NUM_EURO, NUM_EURO, NUM_EURO, NUM_EURO]
        ws.row_dimensions[row].height = 16
        for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=row, column=col, value=val)
            if fill:
                c.fill = _fill(fill)
            c.font = FONT_BOLD if bold else FONT_NORMAL
            c.border = BORDER_THIN
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        row += 1

    def sep_row(label="", color="D6DCE4"):
        nonlocal row
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = FONT_SUBHEADER
        c.fill = _fill(color)
        c.border = BORDER_THIN
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    def empty_row():
        nonlocal row
        ws.row_dimensions[row].height = 8
        row += 1

    # ── INTESTAZIONE ─────────────────────────────────────────
    title_row("ALLEGATO 3", size=9, height=14)
    title_row(
        "MODELLO DEL PIANO ANNUALE DEI FLUSSI DI CASSA — CONTABILITÀ FINANZIARIA",
        size=11, height=24
    )
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value=f"Esercizio finanziario {anno} — {nome}")
    c.font = Font(name="Calibri", bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1
    empty_row()

    header_cols()

    # ── FONDO CASSA INIZIALE ──────────────────────────────────
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=f"FONDO DI CASSA ALL\'INIZIO DELL\'ANNO")
    c.font = FONT_BOLD
    c.fill = _fill(COLOR_HEADER_LIGHT)
    c.border = BORDER_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col, val in enumerate([fondo_cassa, fondo_cassa, 0.0, fondo_cassa], 3):
        c2 = ws.cell(row=row, column=col, value=val)
        c2.font = FONT_BOLD
        c2.fill = _fill(COLOR_HEADER_LIGHT)
        c2.border = BORDER_THIN
        c2.number_format = NUM_EURO
        c2.alignment = Alignment(horizontal="right", vertical="center")
    row += 1
    empty_row()

    # ── SEZIONE RISCOSSIONI ───────────────────────────────────
    title_row("RISCOSSIONI (in c/competenza e in c/residui)", color=COLOR_HEADER_BLUE, height=18)

    # Entrate per codice PDC
    entrate_mim = [
        ("E.1.00.00.00.000", "Entrate correnti di natura tributaria, contributiva e perequativa"),
        ("E.2.00.00.00.000", "Trasferimenti correnti"),
        ("E.2.01.01.00.000", "  Trasferimenti da Amministrazioni pubbliche (Stato, Regione, Enti Locali)"),
        ("E.2.01.02.00.000", "  Trasferimenti correnti da Famiglie (contributi)"),
        ("E.2.01.03.00.000", "  Trasferimenti correnti da Imprese"),
        ("E.2.01.04.00.000", "  Trasferimenti correnti da Istituzioni Sociali Private"),
        ("E.2.01.05.00.000", "  Trasferimenti dall\'Unione Europea (FSE, FESR, altri UE)"),
        ("E.3.00.00.00.000", "Entrate extratributarie"),
        ("E.3.03.00.00.000", "  Interessi attivi"),
        ("E.3.05.00.00.000", "  Rimborsi e altre entrate correnti"),
        ("E.4.00.00.00.000", "Entrate in conto capitale"),
        ("E.5.00.00.00.000", "Entrate da riduzione di attività finanziarie"),
        ("E.6.00.00.00.000", "Accensione Prestiti"),
        ("E.9.00.00.00.000", "Entrate per conto terzi e partite di giro"),
        ("E.9.01.00.00.000", "  Entrate per partite di giro (incl. Fondo Minute Spese)"),
    ]

    # Raggruppa per prefisso PDC
    # Le chiavi nel dict sono brevi (es. "U.1.03", "E.2.01.01")
    # I codici MIM sono lunghi (es. "U.1.03.00.00.000")
    # → confrontiamo il codice MIM con startswith della chiave breve
    def get_pdc_totale(codice_mim):
        ga = sum(v["ga"] for k, v in pdc.items() if codice_mim.startswith(k))
        sd = sum(v["sd"] for k, v in pdc.items() if codice_mim.startswith(k))
        return ga, sd

    tot_risc_ga = ent_ga          # minute spese entrano in Set-Dic (ricostituzione)
    tot_risc_sd = ent_sd + minute_spese

    # Calcola tutti i valori prima di scrivere (serve per sommare le righe padre)
    righe_entrate = []
    calcolati_e = {}  # codice → (ga, sd)
    for codice, desc in entrate_mim:
        ga, sd = get_pdc_totale(codice)
        if codice.startswith("E.9.01"):
            sd += minute_spese
        calcolati_e[codice] = (ga, sd)

    # Calcola righe padre come somma dei figli
    def _somma_discendenti(padre, tutti):
        """Somma tutti i discendenti di una riga padre nel piano flussi MIM."""
        p = padre.replace(".00.00.000","").replace(".00.000","").replace(".000","").replace(".00","")
        ga = sd = 0.0
        for c, (g, s) in tutti.items():
            if c == padre: continue
            cn = c.replace(".00.00.000","").replace(".00.000","").replace(".000","").replace(".00","")
            if cn.startswith(p + "."):
                ga += g; sd += s
        return ga, sd

    for codice, desc in entrate_mim:
        indent = 1 if desc.startswith("  ") else 0
        if indent == 0:
            ga, sd = _somma_discendenti(codice, calcolati_e)
            if ga == 0 and sd == 0:
                ga, sd = calcolati_e[codice]
        else:
            ga, sd = calcolati_e[codice]
        totale = ga + sd
        al_31dic = totale
        desc_clean = desc.strip()
        fill = COLOR_ENTRATE if indent == 0 else None
        data_row_mim(codice, desc_clean, totale, ga, sd, al_31dic,
                     fill=fill, bold=(indent == 0), indent=indent)

    empty_row()

    # Totale riscossioni
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value="TOTALE RISCOSSIONI")
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    c.fill = _fill(COLOR_HEADER_BLUE)
    c.border = BORDER_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col, val in enumerate([tot_risc_ga + tot_risc_sd, tot_risc_ga, tot_risc_sd, tot_risc_ga + tot_risc_sd], 3):
        c2 = ws.cell(row=row, column=col, value=val)
        c2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c2.fill = _fill(COLOR_HEADER_BLUE)
        c2.border = BORDER_THIN
        c2.number_format = NUM_EURO
        c2.alignment = Alignment(horizontal="right", vertical="center")
    row += 1

    # Totale risorse disponibili
    risc_tot = tot_risc_ga + tot_risc_sd
    risorse_ga = fondo_cassa + tot_risc_ga
    risorse_sd = tot_risc_sd
    risorse_tot = fondo_cassa + risc_tot

    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value="TOTALE RISORSE DISPONIBILI (Fondo cassa + Riscossioni)")
    c.font = FONT_BOLD
    c.fill = _fill(COLOR_HEADER_LIGHT)
    c.border = BORDER_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col, val in enumerate([risorse_tot, risorse_ga, risorse_sd, risorse_tot], 3):
        c2 = ws.cell(row=row, column=col, value=val)
        c2.font = FONT_BOLD
        c2.fill = _fill(COLOR_HEADER_LIGHT)
        c2.border = BORDER_THIN
        c2.number_format = NUM_EURO
        c2.alignment = Alignment(horizontal="right", vertical="center")
    row += 1
    empty_row()

    # ── SEZIONE PAGAMENTI ─────────────────────────────────────
    title_row("PAGAMENTI", color="C00000", height=18)

    spese_mim = [
        ("U.1.00.00.00.000", "Spese correnti"),
        ("U.1.01.00.00.000", "  Redditi da lavoro dipendente"),
        ("U.1.02.00.00.000", "  Imposte e tasse a carico dell\'ente"),
        ("U.1.03.00.00.000", "  Acquisto di beni e servizi (Attività A01–A06)"),
        ("U.1.04.00.00.000", "  Trasferimenti correnti"),
        ("U.1.10.00.00.000", "  Altre spese correnti (Progetti P01–P05)"),
        ("U.1.99.00.00.000", "  Fondo di riserva e altre spese"),
        ("U.2.00.00.00.000", "Spese in conto capitale"),
        ("U.2.02.00.00.000", "  Investimenti fissi lordi e acquisto di terreni"),
        ("U.3.00.00.00.000", "Spese per incremento attività finanziarie"),
        ("U.4.00.00.00.000", "Rimborso Prestiti"),
        ("U.5.00.00.00.000", "Chiusura Anticipazioni istituto tesoriere"),
        ("U.7.00.00.00.000", "Uscite per conto terzi e partite di giro"),
        ("U.7.01.00.00.000", "  Uscite per partite di giro (incl. Fondo Minute Spese)"),
    ]

    tot_pag_ga = spe_ga + minute_spese   # minute spese in Gen-Ago (utilizzo)
    tot_pag_sd = spe_sd

    calcolati_s = {}
    for codice, desc in spese_mim:
        ga, sd = get_pdc_totale(codice)
        if codice.startswith("U.7.01"):
            ga += minute_spese
        calcolati_s[codice] = (ga, sd)

    for codice, desc in spese_mim:
        indent = 1 if desc.startswith("  ") else 0
        if indent == 0:
            ga, sd = _somma_discendenti(codice, calcolati_s)
            if ga == 0 and sd == 0:
                ga, sd = calcolati_s[codice]
        else:
            ga, sd = calcolati_s[codice]
        totale = ga + sd
        al_31dic = totale
        desc_clean = desc.strip()
        fill = COLOR_SPESE if indent == 0 else None
        data_row_mim(codice, desc_clean, totale, ga, sd, al_31dic,
                     fill=fill, bold=(indent == 0), indent=indent)

    empty_row()

    # Totale pagamenti
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value="TOTALE PAGAMENTI")
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    c.fill = _fill("C00000")
    c.border = BORDER_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col, val in enumerate([tot_pag_ga + tot_pag_sd, tot_pag_ga, tot_pag_sd, tot_pag_ga + tot_pag_sd], 3):
        c2 = ws.cell(row=row, column=col, value=val)
        c2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c2.fill = _fill("C00000")
        c2.border = BORDER_THIN
        c2.number_format = NUM_EURO
        c2.alignment = Alignment(horizontal="right", vertical="center")
    row += 1
    empty_row()

    # ── SALDI STIMATI ─────────────────────────────────────────
    title_row("SALDO DI CASSA STIMATO", color=COLOR_PIANO_HEADER, height=18)

    saldo_agosto = fondo_cassa + tot_risc_ga - tot_pag_ga
    saldo_dicembre = saldo_agosto + tot_risc_sd - tot_pag_sd

    def _saldo_row(label, totale_12, col_d, col_e, col_f, fill=COLOR_HEADER_LIGHT):
        nonlocal row
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = FONT_BOLD
        c.fill = _fill(fill)
        c.border = BORDER_THIN
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col, val in enumerate([totale_12, col_d, col_e, col_f], 3):
            c2 = ws.cell(row=row, column=col, value=val)
            c2.font = FONT_BOLD
            c2.fill = _fill(fill)
            c2.border = BORDER_THIN
            if isinstance(val, (int, float)):
                c2.number_format = NUM_EURO
                c2.alignment = Alignment(horizontal="right", vertical="center")
                # Colora in rosso se negativo
                if val < 0:
                    c2.font = Font(name="Calibri", bold=True, color="C00000", size=10)
            else:
                c2.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    # Saldo al 31/08: fondo_cassa + riscossioni_gen_ago - pagamenti_gen_ago
    # Col C = saldo_agosto (è il totale cumulato al 31/08)
    # Col D = saldo_agosto (valore al 31/08)
    # Col E = "-" (non applicabile per la colonna "di cui set-dic")
    # Col F = saldo_agosto (al 31/12 parziale = al 31/08)
    _saldo_row(
        f"Fondo di cassa — saldo stimato al 31/08/{anno}",
        totale_12=saldo_agosto,
        col_d=saldo_agosto,
        col_e="-",
        col_f=saldo_agosto,
    )

    # Saldo al 31/12: saldo_agosto + riscossioni_set_dic - pagamenti_set_dic
    # Il fondo cassa al 31/08 DIVENTA il saldo iniziale del periodo Set-Dic
    # Col C = saldo_dicembre
    # Col D = saldo_agosto (il punto di partenza del periodo Set-Dic)
    # Col E = variazione set-dic (risc_sd - pag_sd)
    # Col F = saldo_dicembre (cumulato finale)
    variazione_sd = tot_risc_sd - tot_pag_sd
    _saldo_row(
        f"Fondo di cassa — saldo stimato al 31/12/{anno}",
        totale_12=saldo_dicembre,
        col_d=saldo_agosto,        # saldo iniziale Set-Dic = saldo al 31/08
        col_e=variazione_sd,       # variazione nel periodo Set-Dic
        col_f=saldo_dicembre,      # saldo finale cumulato
    )
    row += 0  # già incrementato da _saldo_row

    if saldo_agosto < 0 or saldo_dicembre < 0:
        empty_row()
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1,
                    value="⚠️ ATTENZIONE: uno o più saldi stimati risultano NEGATIVI. Verificare i dati inseriti.")
        c.font = Font(name="Calibri", bold=True, color="C00000", size=10)
        c.fill = _fill("FFFFC0")
        c.border = BORDER_THIN
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

    empty_row()
    ws.merge_cells(f"A{row}:F{row}")
    modalita_label = totali.get("modalita", "automatica").upper()
    c = ws.cell(row=row, column=1,
                value=f"Elaborato con modalità {modalita_label} · "
                      f"Piano Annuale dei Flussi di Cassa · Art. 6 D.L. 155/2024 · "
                      f"Web App Maurizio Torre")
    c.font = Font(name="Calibri", italic=True, size=9, color="808080")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 14

# Mappatura codice voce H → prefisso PDC per piano flussi
# Entrate: E01..E09 → E.2.x ; Spese: A01..A06 → U.1 ; P01..P05 → U.1 ; R98 → U.1.99
_MAP_ENTRATE_PDC = {
    # Nomenclatura codici E (standard)
    "E01": "E.2.01.01", "E02": "E.2.01.02", "E03": "E.2.01.03",
    "E04": "E.2.01.04", "E05": "E.2.01.05", "E06": "E.3.05",
    "E07": "E.3.05",    "E08": "E.4.00",    "E09": "E.5.00",
    "E10": "E.6.00",    "E11": "E.3.03",    "E99": "E.3.05",
    # Nomenclatura numerica tipo "03/01" (MIM / SIDI) — primi 2 cifre
    "01": "E.1.00",     # Entrate tributarie
    "02": "E.1.00",
    "03": "E.2.01.01",  # Trasferimenti Stato/MIUR
    "04": "E.2.01.01",
    "05": "E.2.01.01",
    "06": "E.2.01.02",  # Contributi famiglie
    "07": "E.2.01.03",  # Contributi imprese
    "08": "E.2.01.05",  # UE / PNRR
    "09": "E.2.01.04",  # Istituzioni sociali private
    "10": "E.3.05",     # Entrate proprie / rimborsi
    "11": "E.3.05",
    "12": "E.3.03",     # Interessi attivi
    "13": "E.3.05",
    "14": "E.4.00",     # Entrate conto capitale
    "15": "E.5.00",
    "99": "E.3.05",     # Altre entrate
}
_MAP_SPESE_PDC = {
    "A01": "U.1.03", "A02": "U.1.03", "A03": "U.1.03",
    "A04": "U.1.03", "A05": "U.1.03", "A06": "U.1.03",
    "P01": "U.1.10", "P02": "U.1.10", "P03": "U.1.10",
    "P04": "U.1.10", "P05": "U.1.10",
    "R98": "U.1.99",
}


def _get_pdc_from_codice(codice: str, tipo: str) -> str:
    """Ricava il prefisso PDC dal codice voce se codice_pdc non disponibile.
    Gestisce sia formato E01/A01 che formato numerico 03/01."""
    if not codice:
        return ""
    cod = codice.strip().upper()
    if tipo not in ("entrata_comp", "residuo_attivo"):
        # Spese: A01, P01, R98 etc — prendi prime 3 lettere
        key = cod.replace("/","").replace(".","").replace("-","")[:3]
        return _MAP_SPESE_PDC.get(key, "U.1.10")
    # Entrate: può essere "E01/01", "03/01", "12/02" etc
    # Prova prima con E-prefix (es E01)
    key3 = cod.replace("/","").replace(".","").replace("-","")[:3]
    if key3 in _MAP_ENTRATE_PDC:
        return _MAP_ENTRATE_PDC[key3]
    # Formato numerico: prendi prime 2 cifre
    import re as _re
    m = _re.match(r"^(\d{2})", cod)
    if m:
        key2 = m.group(1)
        return _MAP_ENTRATE_PDC.get(key2, "E.3.05")
    return "E.3.05"


def genera_excel(dati, dati_scuola, percentuali, modalita) -> bytes:
    """Genera il file Excel completo e lo restituisce come bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    minute_spese = float(dati_scuola.get("fondo_minute_spese", 0))

    tot_entrate = crea_foglio_entrate(wb, dati, percentuali, modalita, minute_spese)
    tot_spese = crea_foglio_spese(wb, dati, percentuali, modalita, minute_spese)

    # Build PDC aggregation for piano flussi
    pdc = {}
    def add_pdc(code, ga, sd):
        if not code:
            return
        if code not in pdc:
            pdc[code] = {"ga": 0.0, "sd": 0.0}
        pdc[code]["ga"] += ga
        pdc[code]["sd"] += sd

    for voce in dati.get("entrate", []):
        codice = voce.get("codice", "")
        if codice in ["E1/1", "E1/2"]:
            continue
        # Usa codice_pdc se disponibile, altrimenti mappa dal codice voce
        code = voce.get("codice_pdc", "") or _get_pdc_from_codice(codice, "entrata_comp")
        if not code or code.upper() == "ESCLUSO":
            continue
        ga, sd, _ = calcola_flussi(voce, percentuali, "entrata_comp", modalita)
        add_pdc(code, ga, sd)

    for residuo in dati.get("residui_attivi", []):
        code = residuo.get("codice_pdc", "") or _get_pdc_from_codice(
            residuo.get("codice", ""), "residuo_attivo")
        if not code:
            code = "E.2.01.01"  # fallback generico entrate trasferimenti
        ga, sd, _ = calcola_flussi(residuo, percentuali, "residuo_attivo", modalita)
        add_pdc(code, ga, sd)

    for voce in dati.get("spese", []):
        codice = voce.get("codice", "")
        code = voce.get("codice_pdc", "") or _get_pdc_from_codice(codice, "spesa_comp")
        if not code:
            code = "U.1.03"  # fallback generico spese correnti
        ga, sd, _ = calcola_flussi(voce, percentuali, "spesa_comp", modalita)
        add_pdc(code, ga, sd)

    for residuo in dati.get("residui_passivi", []):
        code = residuo.get("codice_pdc", "") or _get_pdc_from_codice(
            residuo.get("codice", ""), "residuo_passivo")
        if not code:
            code = "U.1.03"  # fallback generico spese correnti
        ga, sd, _ = calcola_flussi(residuo, percentuali, "residuo_passivo", modalita)
        add_pdc(code, ga, sd)

    totali = {**tot_entrate, **tot_spese, "modalita": modalita, "pdc": pdc}
    fondo_cassa = float(dati_scuola.get("fondo_cassa", 0))

    crea_foglio_piano_flussi(wb, dati_scuola, totali, fondo_cassa)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
